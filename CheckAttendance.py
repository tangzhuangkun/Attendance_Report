import openpyxl 
from openpyxl.styles import Border,Side
#import CheckAttendanceGUI

#version: 3.0
#date: 20190801


class Attendance:
	

	def readData(self,filePath):
		#打开本地存储文档
		wb = openpyxl.load_workbook(filePath)
		dataSheet = wb["每日统计"]
		#dataSheet = wb[1]
		return dataSheet
		
	def analyzeData(self,dataSheet):
		maxRow = dataSheet.max_row
		
		#正常的考勤名称目录
		rightAttendanceSet = {'补卡审批通过','正常','出差','外勤','外出','管理员周侃改为正常',None} 
		
		#一个字典用于记录ID和考勤关系
		attendanceInfo = dict()
		#一个集合用户记录部门
		departmentsSet=set()
		
		#监控ID，判断是否换了一个人了
		tempID = '000000'
		
		for i in range(5,maxRow+1):
			workOrRest = dataSheet.cell(i,7).value
			userID = dataSheet.cell(i,5).value
			department = dataSheet.cell(i,2).value
			#如果用户id变化，属于考勤范围之内，不属于小K小D
			if userID != tempID and workOrRest!='不在考勤组' and ('小K小D' not in department):
				#数据结构，1个名字，6个数字记录,一个list用于存储备注,部门名称
				attendanceInfo[userID] = ['',0,0,0,0,0,0,[],'']
				#存储名字
				name = dataSheet.cell(i,1).value
				attendanceInfo[userID][0] = name
				#存储部门名称
				if '总经办' in department:
					department = '总经办'
				elif '研发中心'in department:
					department = '研发中心'
				attendanceInfo[userID][8] = department
				departmentsSet.add(department)
				#替换监控ID
				tempID = userID
			#当天的考勤日期	
			date = dataSheet.cell(i,6).value[3:9]
			
			
			#如果工作日，并在考勤组
			if '-'in workOrRest and ('小K小D' not in department):
			#V3.0更正了此处逻辑，如果变量workOrRest中包含-，则表示属于考勤范围内的员工
				#上午考勤，若考勤不在正常的考勤名称目录中，在备注中体现
				morningAtten = dataSheet.cell(i,9).value
				if morningAtten not in rightAttendanceSet:
					attendanceInfo[userID][7].append(date+'上班 '+morningAtten) 
				#下午考勤，若考勤不在正常的考勤名称目录中，在备注中体现
				afternoonAtten = dataSheet.cell(i,11).value
				if afternoonAtten not in rightAttendanceSet:
					attendanceInfo[userID][7].append(date+'下班 '+afternoonAtten) 	

		return  attendanceInfo,departmentsSet
		
	def setBorder(self,datasheet,row,col):
		#设置边框
		datasheet.cell(row,col).border =Border(left=Side(style='thin',color='FF000000'),right=Side(style='thin',color='FF000000'),top=Side(style='thin',color='FF000000'),bottom=Side(style='thin',color='FF000000'))
		
		
				
	def getYearMonth(self,dataSheet):
		yearMonth = dataSheet.cell(1,1).value
		#print(yearMonth[-10:-3])
		return yearMonth[-10:-3]
		
	def contentStyle(self,datasheet,row,col,hori_position='center',sizeVal=16,boldBool=False):
		#居中
		datasheet.cell(row,col).alignment = openpyxl.styles.Alignment(horizontal=hori_position, vertical="center", wrap_text=True)
		#字体字号
		datasheet.cell(row,col).font = openpyxl.styles.Font(name="宋体", size=sizeVal, italic=False, color="000000", bold=boldBool)
		
	def backgroundCol(self,datasheet,row,col,color='FFE4C4'):
		#填充单元格颜色
		datasheet.cell(row,col).fill = openpyxl.styles.PatternFill(start_color =color, end_color = color, fill_type = 'solid')  
	
	def generalCellStyle(self,datasheet,row,col):
		self.setBorder(datasheet,row,col);
		self.contentStyle(datasheet,row,col);
		self.backgroundCol(datasheet,row,col);
	
	def employeeAttendanceInfo(self,employee,datasheet,attendanceInfo,rowCounter):
		#写入员工的考勤信息
		#写入姓名	
		datasheet.cell(rowCounter,2).value = attendanceInfo[employee][0]
		self.contentStyle(datasheet,rowCounter,2)
		#设置边框
		self.setBorder(datasheet,rowCounter,2)
		
		#写入备注
		datasheet.cell(rowCounter,10).value = ' ；'.join(attendanceInfo[employee][7])
		self.contentStyle(datasheet,rowCounter,10,hori_position='left')
		#设置边框
		self.setBorder(datasheet,rowCounter,10)
		
		'''
		#写入部门	
		datasheet.cell(rowCounter,11).value = attendanceInfo[employee][8]
		self.contentStyle(datasheet,rowCounter,11)
		#设置边框
		self.setBorder(datasheet,rowCounter,1)
		'''
		
		#设置2-8列边框
		self.setBorder(datasheet,rowCounter,3)
		self.setBorder(datasheet,rowCounter,4)
		self.setBorder(datasheet,rowCounter,5)
		self.setBorder(datasheet,rowCounter,6)
		self.setBorder(datasheet,rowCounter,7)
		self.setBorder(datasheet,rowCounter,8)
		self.setBorder(datasheet,rowCounter,9)
	
	def departmentRow(self,datasheet,rowCounter,department):
		#合并部门行单元格
		datasheet.merge_cells('A'+str(rowCounter)+':J'+str(rowCounter))
		#设置高度
		datasheet.row_dimensions[rowCounter].height = 25
		#写入部门名称
		datasheet.cell(rowCounter,1).value = department
		self.contentStyle(datasheet,rowCounter,1,sizeVal=18,boldBool=True)
		#加边框
		for col in range(1,11):
			self.setBorder(datasheet,rowCounter,col)
		
	def counter(self,datasheet,employeeCounter,rowCounter):
		#写入序列号
		datasheet.cell(rowCounter,1).value = employeeCounter
		self.contentStyle(datasheet,rowCounter,1)
		self.setBorder(datasheet,rowCounter,1)
			
	def writeData(self,departmentsSet,attendanceInfo,yearMonth,savePath):
		wb=openpyxl.Workbook()
		wb.create_sheet(index=0, title=yearMonth)	
		datasheet = wb[yearMonth]
		
		
		#合并首行单元格
		datasheet.merge_cells('A1:J1')
		datasheet.cell(1,1).value = '重庆XXXX科技服务有限公司'+yearMonth+'考勤表'
		self.contentStyle(datasheet,1,1,sizeVal=18,boldBool=True)
		#设置高度
		datasheet.row_dimensions[1].height = 40
		#合并部分均设置边框
		for col in range(1,11):
			self.setBorder(datasheet,1,col)
		
		
		datasheet.cell(2,1).value = '序号'
		self.generalCellStyle(datasheet,2,1)
		
		datasheet.cell(2,2).value = '姓名'
		datasheet.column_dimensions['B'].width = 20
		self.generalCellStyle(datasheet,2,2)
		
		datasheet.cell(2,3).value = '迟到'
		self.generalCellStyle(datasheet,2,3)
		
		datasheet.cell(2,4).value = '早退'
		self.generalCellStyle(datasheet,2,4)
		
		datasheet.cell(2,5).value = '旷工'
		self.generalCellStyle(datasheet,2,5)
		
		datasheet.cell(2,6).value = '事假'
		self.generalCellStyle(datasheet,2,6)
		
		datasheet.cell(2,7).value = '病假'
		self.generalCellStyle(datasheet,2,7)
		
		datasheet.cell(2,8).value = '调休'
		self.generalCellStyle(datasheet,2,8)
		
		datasheet.cell(2,9).value = '公假'
		self.generalCellStyle(datasheet,2,9)
		
		datasheet.cell(2,10).value = '备注'
		self.generalCellStyle(datasheet,2,10)
		#设置宽度
		datasheet.column_dimensions['J'].width = 100
		self.generalCellStyle(datasheet,2,10)
		
		'''
		datasheet.cell(2,11).value = '部门'
		#设置宽度
		datasheet.column_dimensions['K'].width = 18
		self.generalCellStyle(datasheet,2,11)
		'''
		
		#遍历set中所有部门，转化为有序的list
		departmentsList = ['']*len(departmentsSet)	
		backwardCounter = -1
		for department in departmentsSet:
			if department == '总经办':
				departmentsList[0]=department
			elif department == '财务中心':
				departmentsList[1]=department
			elif department == '行政中心':
				departmentsList[2]=department
			elif department == '运营中心':
				departmentsList[3]=department
			elif department == '风管中心':
				departmentsList[4]=department
			elif department == '市场中心':
				departmentsList[5]=department	
			elif department == '研发中心':
				departmentsList[6]=department	
			else:
				departmentsList[backwardCounter]=department
				backwardCounter-=1

		
		employeeCounter = 1
		rowCounter = 3
			
		for department in departmentsList:
			#写入部门行
			self.departmentRow(datasheet,rowCounter,department)
			rowCounter+=1
			#遍历所有员工
			for employee in attendanceInfo:
				#如果员工的部门与现在正在检查的部门一致
				if attendanceInfo[employee][8]==department:
					#将员工信息写入行中
					self.employeeAttendanceInfo(employee, datasheet, attendanceInfo, rowCounter)
					#写入序列号
					self.counter(datasheet,employeeCounter, rowCounter)
					rowCounter+=1
					employeeCounter+=1
		
		
		
		#合并倒数第二行
		rowCounter+=1
		lastButOne = 'A'+str(rowCounter)+':'+'J'+str(rowCounter)
		datasheet.merge_cells(lastButOne)
		#设置倒数第二行背景颜色
		self.backgroundCol(datasheet,rowCounter,1,color='90EE90')
		#self.setBorder(datasheet,rowCounter,1)
		#合并部分均设置边框
		for col in range(1,11):
			self.setBorder(datasheet,rowCounter,col)
		datasheet.cell(rowCounter,1).value=None
		datasheet.row_dimensions[rowCounter].height = 30
		rowCounter+=1
		
		

		#合并倒数第一行
		datasheet.merge_cells('A'+str(rowCounter)+':'+'J'+str(rowCounter))
		#合并部分均设置边框
		for col in range(1,10):
			self.setBorder(datasheet,rowCounter,col)
		#设置高度
		datasheet.row_dimensions[rowCounter].height = 50
		#写入倒数第一行
		datasheet.cell(rowCounter,1).value = '审核：                                                        制表人：'
		self.contentStyle(datasheet,rowCounter,1)
		#加边框
		for col in range(1,11):
			self.setBorder(datasheet,rowCounter,col)

		datasheet.freeze_panes = 'A3'	
		#Windows 版本
		#wb.save(savePath+'xlsx')
		
		#Mac 版本
		wb.save(savePath)
		
		
	def runCheckAttendance(self,filePath,savePath):
		dataSheet = self.readData(filePath)
		attendanceInfo,departmentsSet = self.analyzeData(dataSheet)
		yearMonth = self.getYearMonth(dataSheet)
		self.writeData(departmentsSet,attendanceInfo,yearMonth,savePath)
